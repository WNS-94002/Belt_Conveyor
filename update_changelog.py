"""
update_changelog.py
อ่าน git log และสร้าง/อัปเดต changelog.xlsx
รันได้ตรงๆ หรือถูกเรียกจาก git post-commit hook
"""

import subprocess
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter

XLSX_PATH = os.path.join(os.path.dirname(__file__), 'changelog.xlsx')

# ── สีหลัก ──
C_HEADER_BG  = '1E2333'   # header row background
C_HEADER_FG  = 'EEF0F5'   # header row text
C_ACCENT     = 'F07C1F'   # orange accent
C_ROW_ODD    = '181C27'   # odd row bg
C_ROW_EVEN   = '13161E'   # even row bg
C_TEXT_MAIN  = 'EEF0F5'
C_TEXT_DIM   = '8B90A0'
C_GREEN      = '2ECC71'
C_RED        = 'E74C3C'
C_BORDER     = '2A2F3D'

def thin_border():
    s = Side(style='thin', color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def get_git_log():
    """ดึง git log พร้อม numstat ทุก commit"""
    result = subprocess.run(
        ['git', 'log', '--pretty=format:COMMIT|%H|%h|%ai|%an|%s', '--numstat'],
        capture_output=True, text=True, encoding='utf-8',
        cwd=os.path.dirname(__file__)
    )
    commits = []
    current = None
    for line in result.stdout.splitlines():
        if line.startswith('COMMIT|'):
            if current:
                commits.append(current)
            _, full_hash, short_hash, date_str, author, message = line.split('|', 5)
            dt = datetime.fromisoformat(date_str[:19])
            current = {
                'no':       0,
                'datetime': dt,
                'hash':     short_hash,
                'full_hash':full_hash,
                'author':   author,
                'message':  message,
                'files':    [],
                'added':    0,
                'deleted':  0,
            }
        elif line.strip() and current:
            parts = line.split('\t')
            if len(parts) == 3:
                added_str, del_str, fname = parts
                try:
                    current['added']   += int(added_str)
                    current['deleted'] += int(del_str)
                except ValueError:
                    pass
                current['files'].append(fname)
    if current:
        commits.append(current)

    # เรียงเก่า→ใหม่ และกำหนดลำดับ
    commits.reverse()
    for i, c in enumerate(commits, 1):
        c['no'] = i
    return commits

def build_xlsx(commits):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Changelog'
    ws.sheet_view.showGridLines = False

    # ── Tab color ──
    ws.sheet_properties.tabColor = C_ACCENT

    # ── คอลัมน์และความกว้าง ──
    cols = [
        ('ลำดับ',                    5),
        ('วันที่/เวลา',              20),
        ('Commit',                   9),
        ('ผู้แก้ไข',                 16),
        ('รายละเอียดการเปลี่ยนแปลง', 52),
        ('ไฟล์ที่เปลี่ยนแปลง',       38),
        ('เพิ่ม (+)',                 10),
        ('ลบ (-)',                    10),
    ]
    for col_idx, (_, width) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Title row ──
    ws.row_dimensions[1].height = 36
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value    = 'Conveyor Belt Dashboard — Changelog'
    title_cell.font     = Font(name='Arial', size=14, bold=True, color=C_ACCENT)
    title_cell.fill     = PatternFill('solid', fgColor=C_HEADER_BG)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # ── Subtitle row ──
    ws.row_dimensions[2].height = 18
    ws.merge_cells('A2:H2')
    sub = ws['A2']
    sub.value     = f'สร้างอัตโนมัติจาก git history · อัปเดตล่าสุด: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    sub.font      = Font(name='Arial', size=9, color=C_TEXT_DIM)
    sub.fill      = PatternFill('solid', fgColor=C_HEADER_BG)
    sub.alignment = Alignment(horizontal='center', vertical='center')

    # ── Header row (row 3) ──
    ws.row_dimensions[3].height = 24
    for col_idx, (header, _) in enumerate(cols, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = Font(name='Arial', size=10, bold=True, color=C_HEADER_FG)
        cell.fill      = PatternFill('solid', fgColor=C_ACCENT)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin_border()

    # ── Data rows ──
    for row_offset, c in enumerate(commits):
        row = 4 + row_offset
        ws.row_dimensions[row].height = 42
        bg = C_ROW_ODD if row_offset % 2 == 0 else C_ROW_EVEN

        files_str = '\n'.join(c['files']) if c['files'] else '—'
        values = [
            c['no'],
            c['datetime'].strftime('%d/%m/%Y %H:%M'),
            c['hash'],
            c['author'],
            c['message'],
            files_str,
            c['added'],
            c['deleted'],
        ]
        aligns = ['center','center','center','left','left','left','center','center']

        for col_idx, (val, align) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.fill      = PatternFill('solid', fgColor=bg)
            cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
            cell.border    = thin_border()

            # ── font ──
            if col_idx == 1:  # ลำดับ
                cell.font = Font(name='Arial', size=9, bold=True, color=C_ACCENT)
            elif col_idx == 3:  # Commit hash
                cell.font = Font(name='Courier New', size=9, color='A8C4E0')
            elif col_idx == 7:  # เพิ่ม
                cell.font = Font(name='Arial', size=9, bold=True, color=C_GREEN)
            elif col_idx == 8:  # ลบ
                cell.font = Font(name='Arial', size=9, bold=True,
                                 color=C_RED if c['deleted'] > 0 else C_TEXT_DIM)
            else:
                cell.font = Font(name='Arial', size=9, color=C_TEXT_MAIN
                                 if col_idx in (4, 5) else C_TEXT_DIM)

    # ── Summary row ──
    sum_row = 4 + len(commits)
    ws.row_dimensions[sum_row].height = 20
    ws.merge_cells(f'A{sum_row}:F{sum_row}')
    ws[f'A{sum_row}'].value     = f'รวม {len(commits)} commits'
    ws[f'A{sum_row}'].font      = Font(name='Arial', size=9, bold=True, color=C_TEXT_DIM)
    ws[f'A{sum_row}'].fill      = PatternFill('solid', fgColor=C_HEADER_BG)
    ws[f'A{sum_row}'].alignment = Alignment(horizontal='right', vertical='center')

    ws[f'G{sum_row}'].value     = f'=SUM(G4:G{sum_row-1})'
    ws[f'G{sum_row}'].font      = Font(name='Arial', size=9, bold=True, color=C_GREEN)
    ws[f'G{sum_row}'].fill      = PatternFill('solid', fgColor=C_HEADER_BG)
    ws[f'G{sum_row}'].alignment = Alignment(horizontal='center', vertical='center')

    ws[f'H{sum_row}'].value     = f'=SUM(H4:H{sum_row-1})'
    ws[f'H{sum_row}'].font      = Font(name='Arial', size=9, bold=True, color=C_RED)
    ws[f'H{sum_row}'].fill      = PatternFill('solid', fgColor=C_HEADER_BG)
    ws[f'H{sum_row}'].alignment = Alignment(horizontal='center', vertical='center')

    # ── Freeze panes (ตรึง header) ──
    ws.freeze_panes = 'A4'

    # ── Auto filter ──
    ws.auto_filter.ref = f'A3:H{3 + len(commits)}'

    wb.save(XLSX_PATH)
    print(f'OK: changelog.xlsx ({len(commits)} commits) -> {XLSX_PATH}')

if __name__ == '__main__':
    commits = get_git_log()
    build_xlsx(commits)
